from datetime import datetime
from io import BytesIO

from django.contrib.gis.geos import Point
from django.http import HttpResponse
from django.utils import timezone
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl import Workbook, load_workbook
from rest_framework import filters, status
from rest_framework.decorators import action
from rest_framework.parsers import FormParser, JSONParser, MultiPartParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.viewsets import ModelViewSet

from event_manager.filters import EventFilter
from event_manager.models.events import Event, EventStatus
from event_manager.models.places import Place
from event_manager.serializers.events import EventSerializer


class EventViewSet(ModelViewSet):
    serializer_class = EventSerializer
    parser_classes = (JSONParser, FormParser, MultiPartParser)
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filterset_class = EventFilter
    search_fields = ("name", "place__name")
    ordering_fields = ("name", "starts_at", "ends_at")
    permission_classes = (IsAuthenticated,)


    def get_queryset(self):
        queryset = Event.objects.select_related("author", "place").prefetch_related("images")
        user = self.request.user
        if user.is_superuser:
            return queryset
        return queryset.filter(status=EventStatus.PUBLISHED)

    @action(detail=False, methods=["post"], url_path="import-xlsx")
    def import_xlsx(self, request):
        file = request.FILES.get("file")
        if not file:
            return Response({"detail": "Файл обязателен"}, status=status.HTTP_400_BAD_REQUEST)

        workbook = load_workbook(file)
        sheet = workbook.active
        imported = 0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            (
                name,
                description,
                publish_at,
                starts_at,
                ends_at,
                place_name,
                coords,
                rating,
            ) = row
            lat_str, lon_str = str(coords).split(",")
            place, _ = Place.objects.get_or_create(
                name=place_name,
                defaults={"geo_location": Point(float(lon_str), float(lat_str), srid=4326)},
            )
            Event.objects.create(
                name=name,
                description=description or "",
                publish_at=parse_dt(publish_at),
                starts_at=parse_dt(starts_at),
                ends_at=parse_dt(ends_at),
                place=place,
                rating=int(rating),
                author=request.user,
                status=EventStatus.DRAFT,
            )
            imported += 1
        return Response({"imported": imported})

    @action(detail=False, methods=["get"], url_path="export-xlsx")
    def export_xlsx(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "events"
        sheet.append(
            [
                "Название",
                "Описание",
                "Дата публикации",
                "Дата начала",
                "Дата завершения",
                "Место",
                "Координаты",
                "Рейтинг",
            ]
        )
        for event in queryset:
            sheet.append(
                [
                    event.name,
                    event.description,
                    event.publish_at.isoformat(),
                    event.starts_at.isoformat(),
                    event.ends_at.isoformat(),
                    event.place.name,
                    f"{event.place.geo_location.y},{event.place.geo_location.x}",
                    event.rating,
                ]
            )

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)
        response = HttpResponse(
            stream.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = 'attachment; filename="events.xlsx"'
        return response


def parse_dt(value):
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            return value
        return timezone.make_aware(value)
    parsed = datetime.fromisoformat(str(value))
    if timezone.is_aware(parsed):
        return parsed
    return timezone.make_aware(parsed)
